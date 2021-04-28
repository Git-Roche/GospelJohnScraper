

from typing import Text
import settings as s
import requests
from file_helper import get_data
from bs4 import BeautifulSoup
from os.path import join
from copy import copy
from openpyxl import Workbook

def get_chapter_html():
    """
    Get raw html repsonse for every chapter.

    ### Returns

    list of html text
    """
    chapter_html = []
    for i in range(1, s.chapters + 1):

        fpath = join(s.dl_folder,'chapter{}.pkl'.format(i))
        resp = get_data(
            requests.get,
            fpath,
            url = s.site.format(i)
        )
        chapter_html.append(resp.text)

    return chapter_html

def convert_to_dict(chapter_html):
    """
    Convert html to a dictionary.
    """
    ch_dict = {}
    for i, ch in enumerate(chapter_html):
        soup = BeautifulSoup(ch, 'html.parser')
        tables = soup.find_all('table')
        
        ln_dict = {}
        
        for t in tables:
            data = []
            tname = t.attrs['id']
            tbody = t.find('tbody')
            rows = tbody.find_all('tr')
            for r in rows:
                cols = r.find_all('td')
                cols = [ele.text.strip() for ele in cols]
                data.append([ele for ele in cols if ele])

            ln_dict[tname] = copy(data)

        ch_dict['chapter-{}'.format(i+1)] = copy(ln_dict)

    return ch_dict

def main():
    """
    Scrapes a website that has the English and Greek translations of the gospel of john.
    Saves the result into an excel file.
    """
    
    chapter_html = get_chapter_html()

    ch_dict = get_data(
        convert_to_dict,
        join(s.dl_folder, 'chapter_dict.pkl'),
        chapter_html=chapter_html
    )

    #Construct a list suitable for inserting into an excel spreadsheet.
    chapters = []
    for ch in ch_dict:
        lines = []
        for ln in ch_dict[ch]:
            # print(ch_dict[ch][ln])
            line = {}
            for tr in ch_dict[ch][ln]:

                if len(tr) == 1:
                    #new line
                    line = {
                        'kjv' : tr[0],
                        'greek' : [],
                        'pron' : [],
                        'def' : []
                    }
                elif len(tr) == 2:
                    line = {
                        'kjv' : tr[1],
                        'greek' : [],
                        'pron' : [],
                        'def' : []
                    }

                elif len(tr) == 3:
                    line['greek'].append(tr[1])

                    loc1 = tr[2].find('}')
                    loc11 = tr[2].find('{')
                    loc2 = tr[2].find('\n')
                    line['pron'].append(tr[2][loc11+1:loc1])
                    line['def'].append(tr[2][loc2+1:])
                else:
                    print(tr)
            lines.append(copy(line))
        
        chapter = {
            'Chapter' : ch,
            'Lines' : copy(lines)
        }
        chapters.append(copy(chapter))

    #create excel workbook
    wb = Workbook()
    for ch in chapters:

        sh = wb.create_sheet(ch['Chapter'])
        row_count = 1
        for ln in ch['Lines']:

            sh.cell(row_count, 1, ln['kjv'])
            row_count += 1
            for i, w in enumerate(ln['greek']):
                sh.cell(row_count, i+1, w)
            row_count += 1
            for i, w in enumerate(ln['pron']):
                sh.cell(row_count, i+1, w)
            row_count += 1
            for i, w in enumerate(ln['def']):
                sh.cell(row_count, i+1, w)
            row_count += 2
        
    try:
        wb.remove_sheet('Sheet')
    except:
        pass
    wb.save('Gospel of John.xlsx')
    wb.close()

if __name__ == '__main__':
    main()